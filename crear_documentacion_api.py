from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Función para crear la documentación de APIs en un archivo Excel
def crear_documentacion_api():
    # Crear un nuevo libro de trabajo y hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "API Documentation"

    # Encabezados de la documentación
    headers = [
        "Endpoint", "HTTP Method", "Parameters", "Description", 
        "Auth Required", "Request Example", "Response Example", 
        "Response Code", "Notes"
    ]
    ws.append(headers)  # Añadir encabezados como la primera fila

    # Formatear los encabezados
    for cell in ws[1]:  # ws[1] hace referencia a la primera fila (encabezados)
        cell.font = Font(bold=True, color="FFFFFF")  # Texto en negrita y color blanco
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Color de fondo azul

    # Ejemplo de datos de APIs
    api_data = [
        {
            "Endpoint": "/users",
            "HTTP Method": "GET",
            "Parameters": "id (int) - ID del usuario",
            "Description": "Obtiene la información del usuario",
            "Auth Required": "Sí",
            "Request Example": "GET /users?id=1",
            "Response Example": '{"id": 1, "name": "John Doe", "email": "john@example.com"}',
            "Response Code": 200,
            "Notes": "Endpoint público para obtener datos de usuario"
        },
        {
            "Endpoint": "/users",
            "HTTP Method": "POST",
            "Parameters": "name (str), email (str)",
            "Description": "Crea un nuevo usuario",
            "Auth Required": "Sí",
            "Request Example": 'POST /users { "name": "Jane Smith", "email": "jane@example.com" }',
            "Response Example": '{"id": 2, "name": "Jane Smith", "email": "jane@example.com"}',
            "Response Code": 201,
            "Notes": "Requiere autenticación"
        },
        # Agrega más endpoints aquí
    ]

    # Rellenar la hoja con los datos de API
    for api in api_data:
        ws.append([
            api["Endpoint"], api["HTTP Method"], api["Parameters"], api["Description"], 
            api["Auth Required"], api["Request Example"], api["Response Example"], 
            api["Response Code"], api["Notes"]
        ])

    # Formatear las celdas de datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Color de fondo blanco
            cell.border = cell.border.copy()  # Copiar bordes existentes
            cell.alignment = cell.alignment.copy()  # Copiar alineación existente

    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Obtener la letra de la columna
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar el archivo
    filename = "API_Documentation_Styled.xlsx"
    wb.save(filename)
    print(f"El archivo '{filename}' se ha creado exitosamente.")

# Ejecutar la función
crear_documentacion_api()
